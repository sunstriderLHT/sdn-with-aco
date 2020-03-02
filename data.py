import xlrd
from openpyxl import load_workbook
'''
all parameters we need to predefine before the iteration starts
'''
ant_bw = 1
original_pheromone = 1

# the maximum bandwidth of all links
L_0 = 25

# influence factors
alpha = 2
beta = 1
# evaporate parameter for local update
rau = 0.3
# evaporate parameter for global update
Delta = 0.2

# small as much as possible
delta_tau = 0.3



'''
store the iteration result
they should be immutable once they are returned
also updated according to the result of each iteration
'''

# store iteration result in excel file
def store_iteration_result(file_path, result, iteration):
    sheet_name = f'Sheet{iteration}'
    workbook = load_workbook(file_path)
    new_sheet = workbook.create_sheet(sheet_name)
    for data in result:
        new_sheet.append(data)
    workbook.save(file_path)

def get_iteration_result(file_path):
    result = []
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheets()[-1]
    rows = sheet.nrows
    for row in range(rows):
        result.append(sheet.row_values(row))

    return result



'''
data structure
'''
def get_matrix(path):
    matrix = []
    x_workbook = xlrd.open_workbook(path)
    sheet = x_workbook.sheets()[0]
    rows = sheet.nrows
    for row in range(rows):
        matrix.append(sheet.row_values(row))

    return matrix

# global node state must be updated according to the result of each iteration
def get_global_node_state(file_path):
    global_node_state = []
    x_workbook = xlrd.open_workbook(file_path)
    # get the last iteration result stored in the last sheet
    sheet = x_workbook.sheets()[-1]
    rows = sheet.nrows
    for row in range(rows):
        global_node_state.append(sheet.row_values(row))

    return global_node_state

# def store_global_node_state(file_path, global_node_state):
#     with xlsxwriter.Workbook(file_path) as workbook:
#         worksheet = workbook.add_worksheet()
#         for pheromone, load in enumerate(global_node_state):
#             worksheet.write_row(pheromone, 0, load)

def store_global_node_state(file_path, global_node_state, iteration):
    sheet_name = f'Sheet{iteration}'
    workbook = load_workbook(file_path)
    new_sheet = workbook.create_sheet(sheet_name)

    for node in global_node_state:
        new_sheet.append(node)
    workbook.save(file_path)

# L_list = get_iteration_result('l.xlsx')

topoMatrix = get_matrix('topoMatrix.xlsx')
delay = get_matrix('delay.xlsx')

max_load = [20, 20, 25, 25, 25, 20, 25, 25, 20, 25]

local_node_state = [[original_pheromone, 0] for i in range(10)]
global_node_state = get_global_node_state('globalNodeState.xlsx')

